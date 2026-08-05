#!/usr/bin/env python3
"""示例导出技能：把 AI 分析结果转成 Excel（参考 Claude 导出 Excel 技能思路，可自由改造）。

在应用「导出技能」页配置：
  名称: Excel 导出（示例）
  命令: python
  参数: "D:\TEST Code\AI 文件分析助手\samples\skills\export_excel.py"

脚本会自动收到：
  - 最后一个位置参数 = 分析结果文件 (.md)
  - 环境变量 LITEAI_ANALYSIS_FILE / LITEAI_SOURCE_FILE / LITEAI_OUTPUT_DIR
输出：LITEAI_OUTPUT_DIR / <源文件名>.custom.xlsx
要求：已安装 Python 3 + openpyxl（pip install openpyxl）
"""
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("缺少 openpyxl，请先执行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

analysis = os.environ.get("LITEAI_ANALYSIS_FILE") or sys.argv[-1]
source = os.environ.get("LITEAI_SOURCE_FILE", "unknown")
out_dir = os.environ.get("LITEAI_OUTPUT_DIR", ".")

base = os.path.splitext(os.path.basename(source))[0]
out_path = os.path.join(out_dir, f"{base}.custom.xlsx")

with open(analysis, encoding="utf-8") as f:
    text = f.read()

wb = Workbook()
ws = wb.active
ws.title = "分析结果"
ws.column_dimensions["A"].width = 110

title_font = Font(bold=True, size=14)
heading_font = Font(bold=True, size=11)
heading_fill = PatternFill("solid", fgColor="D9E1F2")
wrap = Alignment(wrap_text=True)

row = 1
ws.cell(row, 1, f"{base} · AI 分析报告").font = title_font
row += 2

for line in text.splitlines():
    s = line.strip()
    if s.startswith("```"):
        row += 1
        continue
    if s.startswith("#"):
        cell = ws.cell(row, 1, s.lstrip("# ").strip())
        cell.font = heading_font
        cell.fill = heading_fill
        row += 1
        continue
    if s == "":
        row += 1
        continue
    cell = ws.cell(row, 1, s)
    cell.alignment = wrap
    row += 1

wb.save(out_path)
print(f"已导出 Excel: {out_path}")
